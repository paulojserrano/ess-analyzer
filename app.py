"""
app.py — Tkinter GUI for the ESS / ASRS Log Analyser.

Run:
    python app.py [optional_path_to.xlsx]

Deps: pandas  numpy  plotly  openpyxl  (tkinter ships with Python)

Architecture
------------
  config.py         — colour constants, ANALYSIS_MODULES registry
  data_loader.py    — Excel sheet detection, build_config()
  analyses/         — one module per domain; each exposes run(data, cfg)
  report_builder.py — generate_html_report() with embedded Plotly charts
  app.py            — tkinter GUI + headless CLI fallback  (this file)
"""
from __future__ import annotations

import os
import sys
import queue as _queue
import threading
import traceback

from data_loader import (
    build_config,
    detect_data_date,
    filter_to_peak_day,
    load_data,
    load_log_day,
    load_user_config,
    validate_data,
    validate_file_path,
)
from config import ANALYSIS_MODULES, AUTO_TYPE_PALETTE, DEFAULT_CHECKED
from analyses import throughput, dwell_time, switch_time, retrieval, fleet_utilization, summary as summary_mod
# from analyses import cycle_time  # temporarily disabled
from report_builder import generate_html_report


# ── analysis pipeline (order = display order in report) ──────────────────────
_PIPELINE = [
    ("throughput", throughput),
    ("dwell",      dwell_time),
    ("switch",     switch_time),
    # ("cycle",      cycle_time),  # temporarily disabled
    ("retrieval",  retrieval),
    ("fleet",      fleet_utilization),
]

# ── preferred chart display order in the HTML report ─────────────────────────
# Charts listed here appear first, in this sequence.
# Any chart IDs not listed are appended afterwards in their original module order.
_CHART_DISPLAY_ORDER = [
    "throughput_total",
    "throughput_heatmap",
    "dwell_heatmap",
    "switch_heatmap",
    "dwell_pick_distribution",
    "throughput_picker_rate",
]


def _sort_registry(registry: list[dict]) -> list[dict]:
    order_map = {cid: i for i, cid in enumerate(_CHART_DISPLAY_ORDER)}
    n = len(_CHART_DISPLAY_ORDER)
    return sorted(registry, key=lambda c: order_map.get(c["id"], n + registry.index(c)))

# ── human-readable step labels ───────────────────────────────────────────────
_STEP_LABEL = {key: lbl for key, lbl in ANALYSIS_MODULES}

# ── file picker (CLI fallback) ────────────────────────────────────────────────

def pick_file() -> str:
    if len(sys.argv) > 1 and os.path.isfile(sys.argv[1]):
        return sys.argv[1]
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk(); root.withdraw()
        path = filedialog.askopenfilename(
            title="Select the ASRS data file",
            filetypes=[
                ("ESS data files", "*.xlsx *.xlsm *.log"),
                ("Excel files", "*.xlsx *.xlsm"),
                ("Log files", "*.log"),
                ("All files", "*.*"),
            ],
        )
        root.destroy()
        return path
    except Exception:
        return input("Enter full path to the .xlsx file: ").strip().strip('"')


# ══════════════════════════════════════════════════════════════════════════════
#  GUI
# ══════════════════════════════════════════════════════════════════════════════

class AnalyzerApp:
    """Tkinter GUI for the ESS / ASRS log analyser."""

    # ── palette ───────────────────────────────────────────────────────────────
    BG     = "#f8fafc"
    CARD   = "#ffffff"
    HEADER = "#0f172a"
    ACCENT = "#e94560"
    TEXT   = "#0f172a"
    MUTED  = "#64748b"
    SUBTLE = "#94a3b8"
    BORDER = "#e2e8f0"
    GREEN  = "#16a34a"
    AMBER  = "#f59e0b"
    BLUE   = "#3b82f6"

    def __init__(self):
        import tkinter as tk
        from tkinter import ttk
        self.tk  = tk
        self.ttk = ttk

        self.root = tk.Tk()
        self.root.title("Hai Robotics ESS Analyzer")
        self.root.geometry("1200x820")
        self.root.minsize(960, 680)
        self.root.configure(bg=self.BG)

        self._q               = _queue.Queue()
        self._days: list[dict] = []   # {"path", "label", "data", "cfg"}
        self._cfg              = None  # cfg from first loaded day
        # Maps iid → list[str] of log file paths for multi-file log days.
        # Single-file loads are also stored here (list of one) for uniform handling.
        self._log_groups: dict[str, list[str]] = {}
        self._outdir           = None
        self._html_path        = None
        self._combined_path    = None
        self._running          = False

        # Loading state — a single counter tracks all in-flight loads across
        # multiple rapid _add_files() calls.  The Run button stays disabled
        # until this reaches zero.
        self._loading_count    = 0
        # Prevents the "ready" status message from being re-printed every
        # time _maybe_enable_run() is called while already ready.
        self._ready_announced  = False

        self._checks = {key: tk.BooleanVar(value=key in DEFAULT_CHECKED) for key, _ in ANALYSIS_MODULES}

        self._setup_styles()
        self._build()
        self.root.after(80, self._poll)

    # ── styles ────────────────────────────────────────────────────────────────

    def _setup_styles(self):
        s = self.ttk.Style(self.root)
        s.theme_use("clam")

        s.configure("TFrame",        background=self.BG)
        s.configure("TLabel",        background=self.BG, foreground=self.TEXT, font=("Segoe UI", 10))
        s.configure("Status.TLabel", background=self.HEADER, foreground=self.SUBTLE, font=("Segoe UI", 8))

        # Primary (dark) button
        s.configure("TButton", background=self.HEADER, foreground="white",
                    font=("Segoe UI", 9), borderwidth=0, relief="flat", padding=(10, 6))
        s.map("TButton",
              background=[("active", "#1e293b"), ("disabled", "#cbd5e1")],
              foreground=[("disabled", "#e2e8f0")])

        # Accent run button
        s.configure("Run.TButton", background=self.ACCENT, foreground="white",
                    font=("Segoe UI", 11, "bold"), borderwidth=0, relief="flat", padding=(10, 10))
        s.map("Run.TButton",
              background=[("active", "#c73652"), ("disabled", "#e2e8f0")],
              foreground=[("disabled", self.SUBTLE)])

        # Ghost / secondary button
        s.configure("Ghost.TButton", background=self.BG, foreground=self.TEXT,
                    font=("Segoe UI", 9), borderwidth=0, relief="flat", padding=(8, 6))
        s.map("Ghost.TButton",
              background=[("active", "#e2e8f0"), ("disabled", self.BG)],
              foreground=[("disabled", self.SUBTLE)])

        # Micro text button for inline toggles
        s.configure("Micro.TButton", background=self.CARD, foreground=self.MUTED,
                    font=("Segoe UI", 8), borderwidth=0, relief="flat", padding=(4, 2))
        s.map("Micro.TButton",
              foreground=[("active", self.TEXT), ("disabled", "#cbd5e1")])

        s.configure("TCheckbutton", background=self.CARD, foreground=self.TEXT,
                    font=("Segoe UI", 9))
        s.map("TCheckbutton", background=[("active", self.CARD)])

        s.configure("Horizontal.TProgressbar",
                    background=self.ACCENT, troughcolor=self.BORDER, borderwidth=0, thickness=4)

        s.configure("Treeview", background=self.CARD, foreground=self.TEXT,
                    fieldbackground=self.CARD, font=("Segoe UI", 9), rowheight=26)
        s.configure("Treeview.Heading", background=self.CARD, foreground=self.SUBTLE,
                    font=("Segoe UI", 7, "bold"), relief="flat")
        s.map("Treeview",
              background=[("selected", "#eff6ff")],
              foreground=[("selected", self.HEADER)])

    # ── layout ────────────────────────────────────────────────────────────────

    def _build(self):
        tk = self.tk

        # ── Header ────────────────────────────────────────────────────────────
        hbar = tk.Frame(self.root, bg=self.HEADER, height=62)
        hbar.pack(fill="x")
        hbar.pack_propagate(False)
        hl = tk.Frame(hbar, bg=self.HEADER)
        hl.pack(side="left", padx=22, fill="y")
        tk.Label(hl, text="Hai Robotics ESS Analyzer",
                 bg=self.HEADER, fg="white",
                 font=("Segoe UI", 15, "bold")).pack(anchor="sw", pady=(16, 0))
        tk.Label(hl, text="Performance analysis from ASRS event-log exports",
                 bg=self.HEADER, fg="#475569",
                 font=("Segoe UI", 8)).pack(anchor="nw", pady=(1, 0))
        # Accent underline
        tk.Frame(self.root, bg=self.ACCENT, height=2).pack(fill="x")

        # ── Status bar (packed before body so it stays at the bottom) ─────────
        sbar = tk.Frame(self.root, bg=self.HEADER, height=26)
        sbar.pack(fill="x", side="bottom")
        sbar.pack_propagate(False)
        self._status_lbl = self.ttk.Label(sbar, text="  No file loaded",
                                           style="Status.TLabel")
        self._status_lbl.pack(side="left", padx=8, pady=4)

        # ── Body (card-based 2-column grid) ───────────────────────────────────
        body = tk.Frame(self.root, bg=self.BG, padx=16, pady=12)
        body.pack(fill="both", expand=True)
        body.columnconfigure(0, weight=1, uniform="half")
        body.columnconfigure(1, weight=1, uniform="half")
        body.rowconfigure(2, weight=1)

        # Row 0: Input Files | Detected Stations
        self._build_files_card(body, 0, 0)
        self._build_stations_card(body, 0, 1)
        # Row 1: Analyses | Run & Output
        self._build_analyses_card(body, 1, 0)
        self._build_run_card(body, 1, 1)
        # Row 2: Analysis Log (full width)
        self._build_log_card(body, 2)

        self.root.bind("<Control-o>", lambda _: self._browse())

    def _make_card(self, parent, row, col, title, colspan=1, right_fn=None):
        """Create a bordered card frame with accent-stripe header. Returns content frame."""
        tk = self.tk
        px = (0, 4) if col == 0 else (4, 0)
        if colspan == 2:
            px = (0, 0)

        outer = tk.Frame(parent, bg=self.BG)
        outer.grid(row=row, column=col, columnspan=colspan, sticky="nsew",
                   padx=px, pady=(0, 8))
        outer.columnconfigure(0, weight=1)
        outer.rowconfigure(1, weight=1)

        card = tk.Frame(outer, bg=self.CARD,
                        highlightbackground=self.BORDER, highlightthickness=1)
        card.pack(fill="both", expand=True)

        # Section header
        sh = tk.Frame(card, bg=self.CARD)
        sh.pack(fill="x", pady=(10, 5))
        tk.Frame(sh, bg=self.ACCENT, width=3, height=12).pack(
            side="left", padx=(12, 0))
        tk.Label(sh, text=title, bg=self.CARD, fg=self.MUTED,
                 font=("Segoe UI", 7, "bold")).pack(side="left", padx=(7, 0))
        if right_fn:
            right_fn(sh)

        # Content area
        content = tk.Frame(card, bg=self.CARD)
        content.pack(fill="both", expand=True, padx=12, pady=(0, 10))
        return content

    # ── Card builders ─────────────────────────────────────────────────────────

    def _build_files_card(self, parent, row, col):
        tk = self.tk

        def _fc_right(sh):
            self._file_count_lbl = tk.Label(
                sh, text="", bg=self.CARD, fg=self.SUBTLE, font=("Segoe UI", 8))
            self._file_count_lbl.pack(side="right", padx=12)

        c = self._make_card(parent, row, col, "INPUT FILES  ·  one per day",
                            right_fn=_fc_right)

        ftf = tk.Frame(c, bg=self.CARD)
        ftf.pack(fill="x", pady=(0, 6))
        self._file_tree = self.ttk.Treeview(
            ftf, columns=("label", "file", "status"), show="headings", height=4)
        self._file_tree.heading("label",  text="Day  ✎")
        self._file_tree.heading("file",   text="Filename")
        self._file_tree.heading("status", text="")
        self._file_tree.column("label",  width=60,  minwidth=50,  anchor="w")
        self._file_tree.column("file",   width=160, minwidth=70,  anchor="w")
        self._file_tree.column("status", width=55,  minwidth=45,  anchor="center")
        self._file_tree.tag_configure("loading", foreground=self.SUBTLE)
        self._file_tree.tag_configure("ready",   foreground=self.GREEN)
        self._file_tree.tag_configure("error",   foreground=self.ACCENT)
        ftsb = self.ttk.Scrollbar(ftf, orient="vertical",
                                   command=self._file_tree.yview)
        self._file_tree.configure(yscrollcommand=ftsb.set)
        self._file_tree.pack(side="left", fill="x", expand=True)
        ftsb.pack(side="right", fill="y")
        self._file_tree.bind("<Double-1>",        self._file_label_edit)
        self._file_tree.bind("<<TreeviewSelect>>", self._on_ft_select)

        fb = tk.Frame(c, bg=self.CARD)
        fb.pack(fill="x", pady=(0, 2))
        self.ttk.Button(fb, text="+  Add files…",
                        command=self._browse).pack(side="left", padx=(0, 4))
        self._remove_btn = self.ttk.Button(fb, text="Remove",
                                            style="Ghost.TButton",
                                            command=self._remove_file,
                                            state="disabled")
        self._remove_btn.pack(side="left")
        tk.Label(fb, text="Ctrl+O", bg=self.CARD, fg=self.SUBTLE,
                 font=("Segoe UI", 8)).pack(side="right", padx=(0, 2))

    def _build_stations_card(self, parent, row, col):
        tk = self.tk

        def _stn_right(sh):
            self._stn_status = tk.Label(
                sh, text="— load a file first", bg=self.CARD,
                fg=self.SUBTLE, font=("Segoe UI", 8))
            self._stn_status.pack(side="right", padx=12)

        c = self._make_card(parent, row, col, "DETECTED STATIONS",
                            right_fn=_stn_right)

        tf = tk.Frame(c, bg=self.CARD)
        tf.pack(fill="x", pady=(0, 4))
        self._tree = self.ttk.Treeview(
            tf, columns=("station", "type", "rate"), show="headings", height=5)
        for col_id, lbl, w, anc in [
            ("station", "Station",    65,  "w"),
            ("type",    "Type  ✎",   150,  "w"),
            ("rate",    "Rate/hr  ✎", 65,  "center"),
        ]:
            self._tree.heading(col_id, text=lbl)
            self._tree.column(col_id, width=w, minwidth=w, anchor=anc)
        tsb = self.ttk.Scrollbar(tf, orient="vertical", command=self._tree.yview)
        self._tree.configure(yscrollcommand=tsb.set)
        self._tree.pack(side="left", fill="x", expand=True)
        tsb.pack(side="right", fill="y")
        self._tree.bind("<Double-1>", self._tree_edit)
        tk.Label(c, text="Double-click Type or Rate/hr to edit",
                 bg=self.CARD, fg=self.SUBTLE,
                 font=("Segoe UI", 8)).pack(anchor="w", pady=(0, 2))

    def _build_analyses_card(self, parent, row, col):
        def _an_right(sh):
            self.ttk.Button(sh, text="None", style="Micro.TButton",
                            command=lambda: self._set_all_checks(False)).pack(
                side="right", padx=(0, 12))
            self.ttk.Button(sh, text="All", style="Micro.TButton",
                            command=lambda: self._set_all_checks(True)).pack(
                side="right", padx=(0, 2))

        c = self._make_card(parent, row, col, "ANALYSES", right_fn=_an_right)

        for key, lbl in ANALYSIS_MODULES:
            self.ttk.Checkbutton(c, text=lbl,
                                 variable=self._checks[key],
                                 style="TCheckbutton").pack(
                anchor="w", padx=8, pady=1)

    def _build_run_card(self, parent, row, col):
        tk = self.tk
        c = self._make_card(parent, row, col, "RUN & OUTPUT")

        self._run_btn = self.ttk.Button(c, text="▶   Run Analysis",
                                         style="Run.TButton",
                                         command=self._run, state="disabled")
        self._run_btn.pack(fill="x", pady=(0, 8))

        self._prog = self.ttk.Progressbar(c, mode="determinate",
                                           style="Horizontal.TProgressbar",
                                           maximum=100)
        self._prog.pack(fill="x", pady=(0, 2))
        self._prog_lbl = tk.Label(c, text="", bg=self.CARD,
                                   fg=self.MUTED, font=("Segoe UI", 8))
        self._prog_lbl.pack(anchor="w", padx=2, pady=(0, 8))

        ob = tk.Frame(c, bg=self.CARD)
        ob.pack(fill="x", pady=(0, 4))
        self._btn_rep = self.ttk.Button(ob, text="Open Report",
                                         style="Ghost.TButton",
                                         command=self._open_report, state="disabled")
        self._btn_rep.pack(side="left", fill="x", expand=True, padx=(0, 4))
        self._btn_dir = self.ttk.Button(ob, text="Open Folder",
                                         style="Ghost.TButton",
                                         command=self._open_folder, state="disabled")
        self._btn_dir.pack(side="left", fill="x", expand=True)

        self._btn_export = self.ttk.Button(c, text="Export All Data (.xlsx)",
                                            style="Ghost.TButton",
                                            command=self._open_combined, state="disabled")
        self._btn_export.pack(fill="x", pady=(0, 4))

    def _build_log_card(self, parent, row):
        tk = self.tk

        def _log_right(sh):
            self.ttk.Button(sh, text="Clear", style="Ghost.TButton",
                            command=self._clear_log).pack(side="right", padx=12)

        c = self._make_card(parent, row, 0, "ANALYSIS LOG", colspan=2,
                            right_fn=_log_right)

        lf = tk.Frame(c, bg=self.CARD)
        lf.pack(fill="both", expand=True)
        lf.columnconfigure(0, weight=1)
        lf.rowconfigure(0, weight=1)

        self._log = tk.Text(lf, wrap="word", bg="#f1f5f9", fg=self.TEXT,
                            font=("Consolas", 9), state="disabled",
                            relief="flat", bd=0, padx=12, pady=12,
                            selectbackground="#dbeafe")
        self._log.grid(row=0, column=0, sticky="nsew")
        lsb = self.ttk.Scrollbar(lf, command=self._log.yview)
        lsb.grid(row=0, column=1, sticky="ns")
        self._log.configure(yscrollcommand=lsb.set)

        self._log.tag_configure("head",  foreground=self.HEADER, font=("Consolas", 9, "bold"))
        self._log.tag_configure("done",  foreground=self.GREEN)
        self._log.tag_configure("fail",  foreground=self.ACCENT)
        self._log.tag_configure("warn",  foreground=self.AMBER)
        self._log.tag_configure("muted", foreground=self.MUTED)
        self._log.tag_configure("step",  foreground=self.BLUE)

        self._log_write("Add one or more .xlsx files to begin.\n", "muted")

    # ── helpers ───────────────────────────────────────────────────────────────

    def _log_write(self, text: str, tag: str | None = None):
        self._log.configure(state="normal")
        self._log.insert("end", text, tag or "")
        self._log.see("end")
        self._log.configure(state="disabled")

    def _clear_log(self):
        self._log.configure(state="normal")
        self._log.delete("1.0", "end")
        self._log.configure(state="disabled")

    def _set_status(self, text: str):
        self._status_lbl.configure(text=f"  {text}")

    # ── Run-button gating ─────────────────────────────────────────────────────

    def _maybe_enable_run(self):
        """Single source of truth for Run button state.

        Enabled only when:  no files are loading  AND  not currently running
                            AND  at least one day has been loaded.
        """
        can_run = (self._loading_count == 0
                   and not self._running
                   and bool(self._days))
        self._run_btn.configure(state="normal" if can_run else "disabled")

        if can_run and not self._ready_announced:
            self._ready_announced = True
            has_lc = any(d["data"].get("lifecycle") is not None for d in self._days)
            n    = len(self._days)
            n_ws = len(self._days[0]["cfg"]["ws_order"]) if self._days else 0
            self._set_status(
                f"Ready — {n} day{'s' if n != 1 else ''}, "
                f"{n_ws} station{'s' if n_ws != 1 else ''} detected"
                + ("" if has_lc else "  |  no lifecycle sheet"))
            if not has_lc:
                self._log_write(
                    "⚠  No lifecycle sheet found — "
                    "Cycle Time and Retrieval will be skipped.\n\n", "warn")

    def _update_file_count(self):
        """Refresh the small file-count badge in the file card header."""
        n = len(self._file_tree.get_children())
        if n == 0:
            self._file_count_lbl.configure(text="")
        elif self._loading_count > 0:
            self._file_count_lbl.configure(
                text=f"{self._loading_count} loading…")
        else:
            self._file_count_lbl.configure(
                text=f"{n} file{'s' if n != 1 else ''}")

    # ── file tree helpers ─────────────────────────────────────────────────────

    def _on_ft_select(self, _=None):
        """Enable Remove only when something is selected in the file tree."""
        has_sel = bool(self._file_tree.selection())
        # Don't allow removing a file that is still loading if analysis is running
        self._remove_btn.configure(state="normal" if has_sel else "disabled")

    def _set_all_checks(self, value: bool):
        for var in self._checks.values():
            var.set(value)

    # ── file management ───────────────────────────────────────────────────────

    def _browse(self):
        from tkinter import filedialog
        paths = filedialog.askopenfilenames(
            title="Select ESS / ASRS data files (one per day)",
            filetypes=[
                ("ESS data files", "*.xlsx *.xlsm *.log"),
                ("Excel files", "*.xlsx *.xlsm"),
                ("Log files", "*.log"),
                ("All files", "*.*"),
            ],
        )
        if paths:
            self._add_files(list(paths))

    def _add_files(self, paths: list[str]):
        from log_converter import extract_log_date

        existing = set(self._file_tree.get_children())

        # ── Group .log files by date so split-day logs merge into one entry ──
        log_by_date: dict[str, list[str]] = {}
        non_log_paths: list[str] = []
        for p in paths:
            if os.path.splitext(p)[1].lower() == ".log":
                date_key = extract_log_date(p) or os.path.basename(p)
                log_by_date.setdefault(date_key, []).append(p)
            else:
                non_log_paths.append(p)

        # Build a unified list of (iid, display_name, group_paths) entries.
        # For log groups the iid is the first (earliest) file in the sorted group.
        # For Excel files the iid is the file path itself.
        entries_to_add: list[tuple[str, str, list[str]]] = []
        for date_key, group in log_by_date.items():
            group_sorted = sorted(group, key=os.path.basename)
            iid = group_sorted[0]
            if iid in existing:
                continue
            n_files = len(group_sorted)
            display = (
                f"{date_key}.log ({n_files} files)" if n_files > 1
                else os.path.basename(iid)
            )
            entries_to_add.append((iid, display, group_sorted))

        for p in non_log_paths:
            if p in existing:
                continue
            entries_to_add.append((p, os.path.basename(p), [p]))

        if not entries_to_add:
            return

        new_paths: list[tuple[str, str, list[str]]] = []  # (iid, display, group)
        for iid, display, group in entries_to_add:
            n = len(existing) + len(new_paths) + 1
            label = f"Day {n}"
            self._file_tree.insert("", "end", iid=iid,
                                   values=(label, display, "Loading…"),
                                   tags=("loading",))
            self._log_groups[iid] = group
            new_paths.append((iid, label, group))

        # Each new logical day increments the loading counter.
        self._loading_count   += len(new_paths)
        self._ready_announced  = False

        self._run_btn.configure(state="disabled")
        self._update_file_count()
        self._set_status(f"Loading {self._loading_count} file{'s' if self._loading_count > 1 else ''}…")

        n = len(new_paths)
        self._log_write(f"Loading {n} day{'s' if n > 1 else ''}…\n\n", "head")

        def _task():
            for iid, label, group in new_paths:
                is_log = os.path.splitext(group[0])[1].lower() == ".log"
                display = self._file_tree.set(iid, "file") if self._file_tree.exists(iid) else os.path.basename(iid)
                self._q.put(("log", (f"  {display}\n", "step")))
                try:
                    # ── File-level validation (first file in group is representative) ─
                    file_vr = validate_file_path(group[0])
                    for w in file_vr.warnings:
                        self._q.put(("log", (f"    ⚠ {w}\n", "warn")))
                    if not file_vr.ok:
                        raise ValueError("; ".join(file_vr.errors))

                    if is_log:
                        n_files = len(group)
                        self._q.put(("log", (
                            f"    Parsing {n_files} log file{'s' if n_files > 1 else ''}…\n",
                            "muted")))
                        data = load_log_day(group)
                    else:
                        self._q.put(("log", ("    Reading Excel sheets…\n", "muted")))
                        data = load_data(iid)

                    p = iid  # keep 'p' name for remainder of block

                    for key, sheet_name in [
                        ("callback",  "callback"),
                        ("station",   "station"),
                        ("lifecycle", "lifecycle"),
                    ]:
                        df = data.get(key)
                        if df is not None:
                            self._q.put(("log", (
                                f"    • {sheet_name}: {len(df):,} rows\n", "muted")))
                        elif key == "lifecycle":
                            self._q.put(("log", (
                                "    • lifecycle: not found "
                                "(cycle/retrieval will be skipped)\n", "warn")))
                        else:
                            self._q.put(("log", (
                                f"    • {sheet_name}: not found\n", "fail")))

                    # ── Data quality validation ──────────────────────────
                    data_vr = validate_data(data)
                    for w in data_vr.warnings:
                        self._q.put(("log", (f"    ⚠ {w}\n", "warn")))
                    if not data_vr.ok:
                        for e in data_vr.errors:
                            self._q.put(("log", (f"    ✗ {e}\n", "fail")))
                        raise ValueError(
                            "Data validation failed — see warnings above."
                        )

                    # ── User config ──────────────────────────────────────
                    user_cfg, cfg_vr = load_user_config(p)
                    for w in cfg_vr.warnings:
                        self._q.put(("log", (f"    ⚠ {w}\n", "warn")))
                    if not cfg_vr.ok:
                        for e in cfg_vr.errors:
                            self._q.put(("log", (f"    ✗ {e}\n", "fail")))
                        raise ValueError(
                            "asrs_config.json validation failed — see errors above."
                        )

                    cfg      = build_config(data, user_cfg)
                    date_lbl = detect_data_date(filter_to_peak_day(data)) or label

                    # ── Duplicate date detection ─────────────────────────
                    existing_dates = {
                        d.get("date_lbl") for d in self._days if d.get("date_lbl")
                    }
                    if date_lbl in existing_dates:
                        self._q.put(("log", (
                            f"    ⚠ Date '{date_lbl}' is already loaded — "
                            f"results may duplicate.\n", "warn")))

                    n_ws    = len(cfg["ws_order"])
                    n_zones = len(set(cfg["type_map"].values()))
                    self._q.put(("log", (
                        f"    • {n_ws} station{'s' if n_ws != 1 else ''}, "
                        f"{n_zones} zone type{'s' if n_zones != 1 else ''}\n", "muted")))
                    if date_lbl and date_lbl != label:
                        self._q.put(("log", (f"    • Date: {date_lbl}\n", "muted")))

                    self._q.put(("day_ready", (p, date_lbl, data, cfg)))
                    self._q.put(("log", ("    ✓ Ready\n\n", "done")))
                except Exception as e:
                    self._q.put(("log", (f"    ✗ {e}\n\n", "fail")))
                    self._q.put(("day_err", (p, str(e))))

        threading.Thread(target=_task, daemon=True).start()

    def _remove_file(self):
        for iid in self._file_tree.selection():
            # If the file is still loading, its loader thread will still emit
            # day_ready / day_err — those handlers will decrement _loading_count
            # and skip the tree update since the row will be gone.
            self._file_tree.delete(iid)
            self._days = [d for d in self._days if d["path"] != iid]
            self._log_groups.pop(iid, None)
        self._remove_btn.configure(state="disabled")
        self._update_file_count()
        if not self._days:
            self._run_btn.configure(state="disabled")
            self._tree_clear()
            self._stn_status.configure(text="— load a file first")
            self._cfg             = None
            self._ready_announced = False
        else:
            self._maybe_enable_run()

    def _file_label_edit(self, event):
        item = self._file_tree.identify_row(event.y)
        col  = self._file_tree.identify_column(event.x)
        if not item or col != "#1":
            return
        bbox = self._file_tree.bbox(item, "#1")
        if not bbox:
            return
        x, y, w, h = bbox
        cur = self._file_tree.set(item, "label")
        ent = self.tk.Entry(self._file_tree, font=("Segoe UI", 9),
                            bg="white", fg=self.TEXT, relief="solid", bd=1)
        ent.place(x=x, y=y, width=w, height=h)
        ent.insert(0, cur)
        ent.select_range(0, "end")
        ent.focus_set()

        def _commit(e=None):
            new_label = ent.get().strip() or cur
            self._file_tree.set(item, "label", new_label)
            for d in self._days:
                if d["path"] == item:
                    d["label"] = new_label
                    break
            ent.destroy()

        ent.bind("<Return>",   _commit)
        ent.bind("<Tab>",      _commit)
        ent.bind("<Escape>",   lambda e: ent.destroy())
        ent.bind("<FocusOut>", _commit)

    # ── station treeview ──────────────────────────────────────────────────────

    def _tree_clear(self):
        for it in self._tree.get_children():
            self._tree.delete(it)

    def _populate_tree(self, cfg: dict):
        self._tree_clear()
        for ws in cfg["ws_order"]:
            typ  = cfg["type_map"].get(ws, "")
            rate = cfg["design_rate"].get(ws, "")
            self._tree.insert("", "end", iid=ws, values=(ws, typ, rate))
        n = len(cfg["ws_order"])
        z = len(set(cfg["type_map"].values()))
        self._stn_status.configure(
            text=f"{n} station{'s' if n != 1 else ''}, {z} zone{'s' if z != 1 else ''}")

    def _tree_edit(self, event):
        item = self._tree.identify_row(event.y)
        col  = self._tree.identify_column(event.x)
        if not item or col not in ("#2", "#3"):
            return
        col_name = ("station", "type", "rate")[int(col[1:]) - 1]
        bbox = self._tree.bbox(item, col)
        if not bbox:
            return
        x, y, w, h = bbox
        cur = self._tree.set(item, col_name)
        ent = self.tk.Entry(self._tree, font=("Segoe UI", 9),
                            bg="white", fg=self.TEXT, relief="solid", bd=1)
        ent.place(x=x, y=y, width=w, height=h)
        ent.insert(0, cur)
        ent.select_range(0, "end")
        ent.focus_set()

        def _commit(e=None):
            self._tree.set(item, col_name, ent.get())
            ent.destroy()

        ent.bind("<Return>",   _commit)
        ent.bind("<Tab>",      _commit)
        ent.bind("<Escape>",   lambda e: ent.destroy())
        ent.bind("<FocusOut>", _commit)

    # ── queue polling ─────────────────────────────────────────────────────────

    def _poll(self):
        try:
            while True:
                kind, payload = self._q.get_nowait()
                if kind == "log":
                    text, tag = payload
                    self._log_write(text, tag)

                elif kind == "progress":
                    val, label = payload
                    self._prog["value"] = val
                    self._prog_lbl.configure(text=label)
                    self._set_status(label)

                elif kind == "day_ready":
                    p, label, data, cfg = payload
                    # Only register if the user hasn't already removed this row
                    if self._file_tree.exists(p):
                        self._days.append({"path": p, "label": label,
                                           "date_lbl": label,
                                           "data": data, "cfg": cfg})
                        self._file_tree.set(p, "label",  label)
                        self._file_tree.set(p, "status", "Ready")
                        self._file_tree.item(p, tags=("ready",))
                        if len(self._days) == 1:
                            self._cfg = cfg
                            self._populate_tree(cfg)
                    # Decrement regardless — the loader thread finished its work
                    self._loading_count = max(0, self._loading_count - 1)
                    self._update_file_count()
                    if self._loading_count > 0:
                        self._set_status(
                            f"Loading… ({self._loading_count} file"
                            f"{'s' if self._loading_count > 1 else ''} remaining)")
                    self._maybe_enable_run()

                elif kind == "day_err":
                    p, _ = payload
                    if self._file_tree.exists(p):
                        self._file_tree.set(p, "status", "Error")
                        self._file_tree.item(p, tags=("error",))
                        self._set_status(f"Load failed: {os.path.basename(p)}")
                    self._loading_count = max(0, self._loading_count - 1)
                    self._update_file_count()
                    self._maybe_enable_run()

                elif kind == "combined_ready":
                    self._combined_path = payload
                    self._btn_export.configure(state="normal")

                elif kind == "done":
                    self._on_done(payload)

        except _queue.Empty:
            pass
        self.root.after(80, self._poll)

    # ── run analysis ──────────────────────────────────────────────────────────

    def _cfg_from_tree(self) -> dict:
        """Rebuild cfg from the current Treeview state (captures user edits)."""
        cfg         = dict(self._cfg)
        type_map    = {}
        design_rate = {}
        for item in self._tree.get_children():
            ws, typ, rate = self._tree.item(item, "values")
            type_map[ws] = typ or self._cfg["type_map"].get(ws, "")
            if rate:
                try:
                    design_rate[ws] = int(rate)
                except ValueError:
                    pass
        unique_types = list(dict.fromkeys(type_map.values()))
        tc = {}
        for i, t in enumerate(unique_types):
            tc[t] = cfg["type_colors"].get(t, AUTO_TYPE_PALETTE[i % len(AUTO_TYPE_PALETTE)])
        cfg["type_map"]          = type_map
        cfg["type_colors"]       = tc
        cfg["design_rate"]       = design_rate
        cfg["design_total_rate"] = sum(design_rate.values()) if design_rate else None
        return cfg

    def _run(self):
        if self._running or not self._days or self._loading_count > 0:
            return

        user_cfg = self._cfg_from_tree()
        outdir   = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), "asrs_analysis_output")
        os.makedirs(outdir, exist_ok=True)
        self._outdir = outdir

        self._run_btn.configure(state="disabled")
        self._btn_rep.configure(state="disabled")
        self._btn_dir.configure(state="disabled")
        self._btn_export.configure(state="disabled")
        self._clear_log()
        self._prog["value"] = 0
        self._prog_lbl.configure(text="")
        self._running = True

        checks = {k: v.get() for k, v in self._checks.items()}
        # Snapshot labels from file tree (user may have renamed)
        days = []
        for iid in self._file_tree.get_children():
            label = self._file_tree.set(iid, "label")
            day   = next((d for d in self._days if d["path"] == iid), None)
            if day:
                days.append({"label": label, "data": day["data"]})

        threading.Thread(
            target=self._bg_run,
            args=(days, user_cfg, outdir, checks),
            daemon=True,
        ).start()

    def _bg_run(self, days: list[dict], user_cfg: dict,
                base_outdir: str, checks: dict[str, bool]):
        """Background thread: run all analyses and generate HTML report."""
        orig_stdout = sys.stdout

        class _Writer:
            def __init__(self, q):
                self._q   = q
                self._buf = ""

            def write(self, text):
                self._buf += text
                while "\n" in self._buf:
                    line, self._buf = self._buf.split("\n", 1)
                    tag = (
                        "done" if "[done]" in line else
                        "fail" if ("[FAIL]" in line or "[ERROR]" in line) else
                        "warn" if ("[warn]" in line or "NOTE" in line) else
                        None
                    )
                    self._q.put(("log", (line + "\n", tag)))

            def flush(self):
                if self._buf:
                    self._q.put(("log", (self._buf, None)))
                    self._buf = ""

        sys.stdout = _Writer(self._q)
        try:
            completed_days: list[dict] = []
            n_days = len(days)

            for day_idx, day in enumerate(days):
                label = day["label"]
                data  = filter_to_peak_day(day["data"])

                safe = "".join(
                    c if c.isalnum() or c in " _-" else "_" for c in label
                ).strip() or f"day_{day_idx + 1}"
                day_outdir = os.path.join(base_outdir, safe)
                os.makedirs(day_outdir, exist_ok=True)

                day_cfg = build_config(data, {})
                day_cfg["type_map"]          = user_cfg["type_map"]
                day_cfg["design_rate"]       = user_cfg["design_rate"]
                day_cfg["design_total_rate"] = user_cfg["design_total_rate"]
                day_cfg["type_colors"]       = user_cfg["type_colors"]

                # "cycle" and "retrieval" require the lifecycle sheet;
                # "fleet" uses both sheets but degrades gracefully if either is absent
                _lifecycle_only = {"cycle", "retrieval"}
                enabled_steps = [
                    (key, mod) for key, mod in _PIPELINE
                    if checks.get(key, False)
                    and (key not in _lifecycle_only or data.get("lifecycle") is not None)
                ]
                skipped_steps = [
                    key for key, _ in _PIPELINE
                    if checks.get(key, False)
                    and key in _lifecycle_only
                    and data.get("lifecycle") is None
                ]
                n_steps = len(enabled_steps)

                day_prefix = f"Day {day_idx + 1}/{n_days}  " if n_days > 1 else ""
                self._q.put(("log", (
                    f"{'─' * 48}\n"
                    f"{day_prefix}{label}\n"
                    f"{'─' * 48}\n", "head")))
                self._q.put(("log", (
                    f"  {n_steps} analysis step{'s' if n_steps != 1 else ''} queued\n",
                    "muted")))
                if skipped_steps:
                    self._q.put(("log", (
                        f"  ⚠ skipped (no lifecycle sheet): "
                        f"{', '.join(_STEP_LABEL.get(k, k) for k in skipped_steps)}\n",
                        "warn")))
                self._q.put(("log", ("\n", None)))

                registry: list[dict] = []

                for step_idx, (key, mod) in enumerate(enabled_steps):
                    step_lbl = _STEP_LABEL.get(key, key)
                    overall  = int(
                        (day_idx / n_days + step_idx / max(n_steps, 1) / n_days) * 95
                    )
                    self._q.put(("progress", (overall, f"{label} — {step_lbl}…")))
                    self._q.put(("log", (f"  → {step_lbl}\n", "step")))
                    try:
                        charts = mod.run(data, day_cfg)
                        registry.extend(charts)
                        self._q.put(("log", (
                            f"    ✓ {len(charts)} chart{'s' if len(charts) != 1 else ''}\n",
                            "done")))
                    except Exception as exc:
                        self._q.put(("log", (
                            f"    ✗ {type(exc).__name__}: {exc}\n", "fail")))
                        traceback.print_exc()

                self._save_excel(data, day_cfg, day_outdir)

                total_charts = len(registry)
                self._q.put(("log", (
                    f"\n  {total_charts} chart{'s' if total_charts != 1 else ''} generated"
                    f"  •  output → {os.path.basename(day_outdir)}/\n\n",
                    "muted")))

                completed_days.append({
                    "label":    label,
                    "outdir":   day_outdir,
                    "registry": _sort_registry(registry),
                    "data":     data,   # filtered data — used by summary analysis
                    "cfg":      day_cfg,
                })

            # Cross-day summary charts (only meaningful with ≥ 2 days)
            summary_registry: list[dict] = []
            if n_days > 1:
                try:
                    all_days_raw = [
                        {"label": d["label"], "data": d["data"], "cfg": d["cfg"]}
                        for d in completed_days
                    ]
                    summary_registry = summary_mod.run(all_days_raw)
                    summary_mod.export_xlsx(all_days_raw, base_outdir)
                except Exception as exc:
                    self._q.put(("log", (f"  ⚠ summary charts skipped: {exc}\n", "warn")))

            try:
                combined_path = self._save_combined_excel(completed_days, base_outdir)
                self._q.put(("combined_ready", combined_path))
            except Exception as exc:
                self._q.put(("log", (f"  ⚠ combined export skipped: {exc}\n", "warn")))

            self._q.put(("progress", (98, "Building HTML report…")))
            self._q.put(("log", ("Building HTML report…\n", "step")))
            html_path   = generate_html_report(base_outdir, completed_days,
                                               summary_registry=summary_registry)
            grand_total = sum(len(d["registry"]) for d in completed_days)
            self._q.put(("progress", (100, f"Done — {grand_total} chart(s)")))
            self._q.put(("log", (
                f"{'─' * 48}\n"
                f"✓  Complete  —  {grand_total} chart{'s' if grand_total != 1 else ''} "
                f"across {n_days} day{'s' if n_days != 1 else ''}\n"
                f"{'─' * 48}\n", "done")))
            self._q.put(("done", html_path))

        finally:
            sys.stdout = orig_stdout

    @staticmethod
    def _save_excel(data: dict, cfg: dict, outdir: str):
        """Save Excel tables that were previously produced by the old analysis functions."""
        import pandas as pd

        cb  = data.get("callback")
        lsr = data.get("station")
        tlc = data.get("lifecycle")

        # Throughput Excel — uses triggerGo events from station record,
        # consistent with the throughput analysis charts.
        if lsr is not None:
            try:
                ws_order = cfg["ws_order"]
                p2ws     = cfg["point2ws"]
                lsr_tp = lsr.copy()
                lsr_tp["ts"]      = pd.to_datetime(lsr_tp["时间戳"])
                lsr_tp["station"] = lsr_tp["位置编号"].map(p2ws)

                amr_type = cfg.get("amr_type")
                if amr_type and "机器人类型" in lsr_tp.columns:
                    lsr_tp = lsr_tp[lsr_tp["机器人类型"] == amr_type]

                tgo = lsr_tp[lsr_tp["事件类型"] == "triggerGo"].dropna(subset=["station"])
                tgo = tgo[tgo["station"].isin(ws_order)]
                tgo["hour"] = tgo["ts"].dt.floor("h")

                pivot = tgo.groupby(["hour", "station"]).size().unstack(fill_value=0)
                order = [w for w in ws_order if w in pivot.columns]
                if order:
                    pivot = pivot[order]
                    out   = pivot.copy()
                    out.index = [h.strftime("%H:00") for h in out.index]
                    out.index.name = "Hour"
                    out["TOTAL"] = out.sum(axis=1)
                    out.loc["TOTAL"] = out.sum()
                    with pd.ExcelWriter(
                        os.path.join(outdir, "throughput_by_workstation_hour.xlsx")
                    ) as w:
                        out.to_excel(w, sheet_name="throughput_by_hour")
                        raw_tp = tgo[["时间戳", "station", "事件类型"]].rename(columns={
                            "时间戳": "Timestamp", "station": "Station", "事件类型": "Event",
                        })
                        raw_tp.to_excel(w, sheet_name="raw_events", index=False)
            except Exception:
                pass

        # Cycle time Excel
        from config import TOTAL_DURATION_COL
        if tlc is not None:
            try:
                tj = tlc.copy()
                tj["complete_ts"] = pd.to_datetime(tj["complete(任务完成时间)"])
                tj = tj.dropna(subset=[TOTAL_DURATION_COL])
                tj = tj[(tj[TOTAL_DURATION_COL] >= 0) & (tj[TOTAL_DURATION_COL] < 7200)]
                tj["cycle_min"] = tj[TOTAL_DURATION_COL] / 60
                with pd.ExcelWriter(
                    os.path.join(outdir, "cycle_time_distribution.xlsx")
                ) as w:
                    tj[[TOTAL_DURATION_COL, "cycle_min"]].to_excel(
                        w, sheet_name="cycle_time_intervals", index=False)
                    tj["cycle_min"].describe(
                        percentiles=[0.5, 0.75, 0.9, 0.95, 0.99]
                    ).round(2).reset_index().to_excel(
                        w, sheet_name="summary_stats", index=False)
                    tj.drop(columns=["complete_ts"], errors="ignore").to_excel(
                        w, sheet_name="raw_lifecycle", index=False)
            except Exception:
                pass

        # Dwell capacity utilisation Excel
        if lsr is not None and cb is not None:
            try:
                ws_order = cfg["ws_order"]
                p2ws     = cfg["point2ws"]

                lsr2 = lsr.copy()
                lsr2["ts"]      = pd.to_datetime(lsr2["时间戳"])
                lsr2["station"] = lsr2["位置编号"].map(p2ws)

                ev2       = lsr2.sort_values(["机器人编号", "ts"])
                pick_rows: list[dict] = []
                for _rb, sub in ev2.groupby("机器人编号"):
                    arr = arr_s = None
                    for ts, et, st in sub[["ts", "事件类型", "station"]].values:
                        if et == "arrived":
                            arr, arr_s = ts, st
                        elif et == "triggerGo" and arr is not None:
                            v = (ts - arr).total_seconds()
                            if 0 <= v < 3600 and pd.notna(arr_s):
                                pick_rows.append({"station": str(arr_s), "hour": arr.floor("h"), "pick_s": v})
                            arr = None

                if pick_rows:
                    from analyses.dwell_time import _clipped_occupancy

                    pf      = pd.DataFrame(pick_rows)
                    hrs     = pd.date_range(pf["hour"].min().normalize(), periods=24, freq="h")
                    avg_pick = (
                        pf.groupby(["station", "hour"])["pick_s"].mean()
                          .unstack().reindex(index=ws_order, columns=hrs)
                    )

                    # Build pick events with timestamps for clipped occupancy
                    pick_ev2: list[dict] = []
                    for _rb, sub in ev2.groupby("机器人编号"):
                        _arr = _arr_s = None
                        for ts, et, st in sub[["ts", "事件类型", "station"]].values:
                            if et == "arrived":
                                _arr, _arr_s = ts, st
                            elif et == "triggerGo" and _arr is not None:
                                v = (ts - _arr).total_seconds()
                                if 0 <= v < 3600 and pd.notna(_arr_s):
                                    pick_ev2.append({"station": str(_arr_s), "arr_ts": _arr, "tg_ts": ts})
                                _arr = None
                    pick_df = pd.DataFrame(pick_ev2)
                    pick_occ = _clipped_occupancy(pick_df, "arr_ts", "tg_ts", ws_order, hrs)
                    util_pick = (pick_occ / 3600.0 * 100.0)

                    ev_sw = lsr2[lsr2["事件类型"].isin(["release", "arrived"]) & lsr2["station"].notna()]
                    ev_sw = ev_sw.sort_values(["station", "ts"])
                    sw_rows: list[dict] = []
                    for ws_sw, sub in ev_sw.groupby("station"):
                        rel = None
                        for ts, et in sub[["ts", "事件类型"]].values:
                            if et == "release":
                                rel = ts
                            elif et == "arrived" and rel is not None:
                                v = (ts - rel).total_seconds()
                                if 0 <= v < 7200:
                                    sw_rows.append({
                                        "station": ws_sw, "hour": rel.floor("h"),
                                        "switch_s": v, "rel_ts": rel, "next_arr_ts": ts,
                                    })
                                rel = None

                    actual = (
                        pf.groupby(["hour", "station"]).size()
                          .unstack(fill_value=0)
                          .reindex(index=hrs, columns=ws_order, fill_value=0)
                          .T
                    )

                    def _fmt(p: pd.DataFrame) -> pd.DataFrame:
                        out = p.copy()
                        out.columns = [c.strftime("%H:00") for c in out.columns]
                        out.index.name = "Station"
                        return out.round(1)

                    with pd.ExcelWriter(os.path.join(outdir, "dwell_capacity_utilisation.xlsx")) as w:
                        _fmt(avg_pick).to_excel(w, sheet_name="avg_pick_time_s")
                        _fmt(actual.reindex(ws_order)).to_excel(w, sheet_name="actual_completions")
                        _fmt(pick_occ.reindex(ws_order)).to_excel(w, sheet_name="pick_seconds_per_hour")
                        _fmt(util_pick.reindex(ws_order)).to_excel(w, sheet_name="pick_occupancy_pct")
                        if sw_rows:
                            sw_df = pd.DataFrame(sw_rows)
                            switch_occ = _clipped_occupancy(
                                sw_df, "rel_ts", "next_arr_ts", ws_order, hrs,
                            )
                            full_occ = (pick_occ + switch_occ).clip(upper=3600.0)
                            util_full = full_occ / 3600.0 * 100.0
                            _fmt(switch_occ.reindex(ws_order)).to_excel(w, sheet_name="switch_seconds_per_hour")
                            _fmt(util_full.reindex(ws_order)).to_excel(w, sheet_name="station_occupancy_pct")
                            pd.DataFrame(sw_rows).drop(columns=["rel_ts", "next_arr_ts"]).rename(columns={
                                "station": "Station", "hour": "Hour", "switch_s": "Switch Time (s)",
                            }).to_excel(w, sheet_name="raw_switch_events", index=False)
                        pd.DataFrame(pick_rows).rename(columns={
                            "station": "Station", "hour": "Hour", "pick_s": "Pick Time (s)",
                        }).to_excel(w, sheet_name="raw_pick_events", index=False)
            except Exception:
                pass

        # Fleet delivery leg Excel
        if tlc is not None:
            try:
                dest_col_f = next((c for c in tlc.columns if "目标位置" in str(c)), None)
                leg_col_f  = next((c for c in tlc.columns if "完成耗时" in str(c)), None)
                if dest_col_f and leg_col_f:
                    df_leg = tlc[tlc[dest_col_f].astype(str).str.startswith("LABOR")].copy()
                    df_leg = df_leg.dropna(subset=[leg_col_f])
                    df_leg = df_leg[(df_leg[leg_col_f] > 0) & (df_leg[leg_col_f] < 3600)]
                    if not df_leg.empty:
                        df_leg["Station"]          = df_leg[dest_col_f].astype(str)
                        df_leg["Delivery Leg (s)"] = df_leg[leg_col_f].astype(float).round(1)
                        ws_order_f = cfg.get("ws_order", sorted(df_leg["Station"].unique()))
                        stats_leg = (
                            df_leg.groupby("Station")["Delivery Leg (s)"]
                                  .describe(percentiles=[0.25, 0.5, 0.75, 0.9])
                                  .round(1)
                                  .reindex(ws_order_f)
                        )
                        with pd.ExcelWriter(os.path.join(outdir, "fleet_delivery_leg.xlsx")) as w:
                            stats_leg.to_excel(w, sheet_name="delivery_leg_stats_by_station")
                            df_leg[["Station", "Delivery Leg (s)"]].to_excel(
                                w, sheet_name="raw_deliveries", index=False
                            )
            except Exception:
                pass

        # Fleet utilisation profile Excel
        if lsr is not None:
            try:
                lsr_u   = lsr.copy()
                lsr_u["ts"] = pd.to_datetime(lsr_u["时间戳"], errors="coerce")
                rb_col  = next((c for c in lsr_u.columns if "机器人编号" in str(c)), None)
                rbt_col = next((c for c in lsr_u.columns if "机器人类型" in str(c)), None)
                if rb_col and rbt_col:
                    lsr_u = lsr_u.dropna(subset=["ts", rb_col, rbt_col])
                    util_rows: list[dict] = []
                    for rtype, fleet_df in lsr_u.groupby(rbt_col):
                        fleet_size = fleet_df[rb_col].nunique()
                        fd = fleet_df.copy()
                        fd["hour"] = fd["ts"].dt.floor("h")
                        for hr, cnt in fd.groupby("hour")[rb_col].nunique().items():
                            util_rows.append({
                                "Hour":            hr.strftime("%H:00"),
                                "Robot Type":      rtype,
                                "Active Robots":   int(cnt),
                                "Fleet Size":      fleet_size,
                                "Utilisation (%)": round(cnt / fleet_size * 100, 1) if fleet_size > 0 else None,
                            })
                    if util_rows:
                        with pd.ExcelWriter(
                            os.path.join(outdir, "fleet_utilization_profile.xlsx")
                        ) as w:
                            pd.DataFrame(util_rows).to_excel(
                                w, sheet_name="utilisation_by_hour", index=False
                            )
                            lsr_u.to_excel(w, sheet_name="raw_station_events", index=False)
            except Exception:
                pass

    @staticmethod
    def _save_combined_excel(completed_days: list[dict], base_outdir: str) -> str:
        """Write a self-auditable workbook: raw data + formula-derived + computed sheets.

        Sheet taxonomy (per day, prefixed D{n}_ when multi-day):
          *_raw           — verbatim source data, no filtering
          tp_events       — LABOR completion events with Hour bucket (intermediate)
          tp_pivot        — throughput matrix using Excel COUNTIFS (formula)
          pick_events     — arrived→triggerGo pairs with Pick_s formula column
          switch_events   — release→arrived pairs with Switch_s formula column
          dwell_derived   — full utilisation formula chain, one row per Station×Hour
          cycle_events    — lifecycle rows within bounds + Cycle_min formula column
          cycle_stats     — descriptive stats (computed)
          delivery_leg    — delivery-leg durations + Station column
          leg_stats       — per-station percentiles (computed)
          fleet_util      — hourly fleet utilisation (computed, complex interval logic)
        """
        import pandas as pd
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        from config import TOTAL_DURATION_COL

        multi    = len(completed_days) > 1
        out_path = os.path.join(base_outdir, "all_data.xlsx")

        # ── openpyxl styling helpers ──────────────────────────────────────────
        _DARK  = "0F172A"
        _BLUE  = "1D4ED8"
        _LBLUE = "EFF6FF"
        _AMBER = "FEF3C7"

        def _style_header(ws):
            """Bold white-on-dark header for row 1."""
            for col in range(1, (ws.max_column or 1) + 1):
                c = ws.cell(1, col)
                if c.value is not None:
                    c.font = Font(bold=True, color="FFFFFF")
                    c.fill = PatternFill("solid", fgColor=_DARK)

        def _note_cell(ws, row, col, text):
            """Write an italicised blue annotation cell."""
            c = ws.cell(row, col, text)
            c.font = Font(italic=True, color=_BLUE)
            c.fill = PatternFill("solid", fgColor=_LBLUE)

        def _method_cell(ws, row, col, text):
            """Write an amber method-explanation cell."""
            c = ws.cell(row, col, text)
            c.font = Font(italic=True, color="92400E")
            c.fill = PatternFill("solid", fgColor=_AMBER)

        def _sn(base: str, prefix: str) -> str:
            name = f"{prefix}{base}" if prefix else base
            return name[:31]

        with pd.ExcelWriter(out_path, engine="openpyxl") as w:
            wb = w.book

            # ── Index / legend sheet ──────────────────────────────────────────
            legend_rows = [
                ["Sheet suffix", "Type", "Description"],
                ["cb_raw",        "Raw",      "Complete callback sheet — all robot events, no filtering"],
                ["st_raw",        "Raw",      "Complete station sheet — all labour-station robot events"],
                ["lc_raw",        "Raw",      "Complete lifecycle sheet — full container journey records"],
                ["tp_events",     "Data",     "TriggerGo events from station record at LABOR stations; adds Hour bucket column"],
                ["tp_pivot",      "Formula",  "Throughput count matrix — cells are live COUNTIFS(tp_events)"],
                ["pick_events",   "Data",     "Arrived→TriggerGo pairs per robot; Pick_s column is Excel formula"],
                ["switch_events", "Data",     "Release→Arrived pairs per station; Switch_s column is Excel formula"],
                ["dwell_derived", "Formula",  "Full utilisation chain: pick_avg, switch_avg, implied TPH, util% — all live formulas referencing pick_events / switch_events / tp_events"],
                ["cycle_events",  "Data",     "Lifecycle rows within valid bounds; Cycle_min column is Excel formula"],
                ["cycle_stats",   "Computed", "Cycle time descriptive statistics (Python-computed percentiles)"],
                ["delivery_leg",  "Data",     "Delivery legs to LABOR stations from lifecycle sheet"],
                ["leg_stats",     "Computed", "Per-station delivery leg percentiles (Python-computed)"],
                ["fleet_util",    "Computed", "Hourly fleet utilisation — active unique robots / fleet size × 100"],
            ]
            legend_df = pd.DataFrame(legend_rows[1:], columns=legend_rows[0])
            legend_df.to_excel(w, sheet_name="_Index", index=False)
            _style_header(wb["_Index"])
            wb["_Index"].column_dimensions["A"].width = 18
            wb["_Index"].column_dimensions["B"].width = 12
            wb["_Index"].column_dimensions["C"].width = 80
            if multi:
                note = wb["_Index"].cell(len(legend_rows) + 2, 1,
                    f"Multi-day run: each sheet is prefixed D1_, D2_, … (D1_ = {completed_days[0]['label']})")
                note.font = Font(italic=True, color=_BLUE)

            # ── Per-day sheets ────────────────────────────────────────────────
            for day_idx, day in enumerate(completed_days):
                data = day["data"]
                cfg  = day["cfg"]
                p    = f"D{day_idx + 1}_" if multi else ""

                cb  = data.get("callback")
                lsr = data.get("station")
                tlc = data.get("lifecycle")
                ws_order = cfg.get("ws_order", [])
                p2ws     = cfg.get("point2ws", {})

                # ── 1. Raw sheets ─────────────────────────────────────────
                for sbase, df in [("cb_raw", cb), ("st_raw", lsr), ("lc_raw", tlc)]:
                    if df is not None:
                        try:
                            df.to_excel(w, sheet_name=_sn(sbase, p), index=False)
                            _style_header(wb[_sn(sbase, p)])
                        except Exception:
                            pass

                # ── 2. Throughput events intermediate sheet ───────────────
                # Uses triggerGo events from the station record — consistent
                # with the throughput analysis charts.
                tp_sn = None
                if lsr is not None:
                    try:
                        lsr_tp = lsr.copy()
                        lsr_tp["ts"]      = pd.to_datetime(lsr_tp["时间戳"])
                        lsr_tp["Station"] = lsr_tp["位置编号"].map(p2ws)

                        amr_type = cfg.get("amr_type")
                        if amr_type and "机器人类型" in lsr_tp.columns:
                            lsr_tp = lsr_tp[lsr_tp["机器人类型"] == amr_type]

                        tgo = lsr_tp[lsr_tp["事件类型"] == "triggerGo"].dropna(subset=["Station"])
                        tgo = tgo[tgo["Station"].isin(ws_order)].copy()
                        tgo["Hour"] = tgo["ts"].dt.strftime("%H:00")
                        tp_sn = _sn("tp_events", p)
                        tgo[["时间戳", "Hour", "Station"]].rename(
                            columns={"时间戳": "Timestamp"}
                        ).to_excel(w, sheet_name=tp_sn, index=False)
                        _style_header(wb[tp_sn])
                        # col D: annotation explaining the Hour bucket
                        _note_cell(wb[tp_sn], 1, 4,
                            "Hour = Timestamp floored to HH:00  |  "
                            "Station = LABOR station mapped from 位置編號  |  "
                            "Filter: 事件類型 == 'triggerGo' at mapped LABOR stations")
                    except Exception:
                        pass

                # ── 3. Throughput formula pivot ───────────────────────────
                if tp_sn and ws_order:
                    try:
                        pivot_sn = _sn("tp_pivot", p)
                        pws      = wb.create_sheet(pivot_sn)
                        n_st     = len(ws_order)
                        hours    = [f"{h:02d}:00" for h in range(24)]
                        total_col = n_st + 2

                        # Row 1 — header
                        pws.cell(1, 1, "Hour")
                        for ci, st in enumerate(ws_order, 2):
                            pws.cell(1, ci, st)
                        pws.cell(1, total_col, "TOTAL")
                        _style_header(pws)

                        # Row 2 — formula legend (blue annotation)
                        _note_cell(pws, 2, 1, "FORMULA →")
                        _note_cell(pws, 2, 2,
                            f"=COUNTIFS('{tp_sn}'!$C:$C,[Station],'{tp_sn}'!$B:$B,[Hour])")
                        for ci in range(3, total_col + 1):
                            pws.cell(2, ci).fill = PatternFill("solid", fgColor=_LBLUE)

                        # Rows 3..26 — data with live COUNTIFS
                        for ri, hr in enumerate(hours, 3):
                            pws.cell(ri, 1, hr)
                            for ci, _ in enumerate(ws_order, 2):
                                scol = get_column_letter(ci)
                                pws.cell(ri, ci,
                                    f"=COUNTIFS('{tp_sn}'!$C:$C,{scol}$1,"
                                    f"'{tp_sn}'!$B:$B,$A{ri})")
                            last_data = get_column_letter(n_st + 1)
                            pws.cell(ri, total_col, f"=SUM(B{ri}:{last_data}{ri})")

                        # TOTAL row
                        tr = 3 + len(hours)
                        pws.cell(tr, 1, "TOTAL")
                        for ci in range(2, total_col + 1):
                            cl = get_column_letter(ci)
                            pws.cell(tr, ci, f"=SUM({cl}3:{cl}{tr - 1})")
                        for ci in range(1, total_col + 1):
                            c = pws.cell(tr, ci)
                            c.font = Font(bold=True)
                    except Exception:
                        pass

                # ── 4. Pick events with Excel-formula Pick_s ──────────────
                pick_sn = None
                if lsr is not None:
                    try:
                        lsr2 = lsr.copy()
                        lsr2["ts"]      = pd.to_datetime(lsr2["时间戳"])
                        lsr2["Station"] = lsr2["位置编号"].map(p2ws)
                        ev2   = lsr2.sort_values(["机器人编号", "ts"])
                        rows  = []
                        for rb, sub in ev2.groupby("机器人编号"):
                            arr_ts = arr_s = None
                            for ts, et, st in sub[["ts", "事件类型", "Station"]].values:
                                if et == "arrived":
                                    arr_ts, arr_s = ts, st
                                elif et == "triggerGo" and arr_ts is not None:
                                    v = (ts - arr_ts).total_seconds()
                                    if 0 <= v < 3600 and pd.notna(arr_s):
                                        rows.append({
                                            "Robot_ID": rb,
                                            "Station":  str(arr_s),
                                            "Hour":     arr_ts.strftime("%H:00"),
                                            "Arrived":  arr_ts,
                                            "TriggerGo": ts,
                                        })
                                    arr_ts = None
                        if rows:
                            pick_sn = _sn("pick_events", p)
                            pd.DataFrame(rows).to_excel(w, sheet_name=pick_sn, index=False)
                            ews = wb[pick_sn]
                            _style_header(ews)
                            # Col F = Pick_s formula; col G = annotation
                            ews.cell(1, 6, "Pick_s")
                            ews.cell(1, 6).font = Font(bold=True, color="FFFFFF")
                            ews.cell(1, 6).fill = PatternFill("solid", fgColor=_DARK)
                            _note_cell(ews, 1, 7,
                                "Pick_s = (TriggerGo − Arrived) × 86400  "
                                "| Event pair: arrived → triggerGo per Robot_ID  "
                                "| Filter: 0 ≤ Pick_s < 3600")
                            for r in range(2, len(rows) + 2):
                                ews.cell(r, 6, f"=(E{r}-D{r})*86400")
                    except Exception:
                        pass

                # ── 5. Switch events with Excel-formula Switch_s ──────────
                switch_sn = None
                if lsr is not None:
                    try:
                        lsr3 = lsr.copy()
                        lsr3["ts"]      = pd.to_datetime(lsr3["时间戳"])
                        lsr3["Station"] = lsr3["位置编号"].map(p2ws)
                        ev_sw = lsr3[
                            lsr3["事件类型"].isin(["release", "arrived"]) &
                            lsr3["Station"].notna()
                        ].sort_values(["Station", "ts"])
                        rows = []
                        for ws_name, sub in ev_sw.groupby("Station"):
                            rel_ts = None
                            for ts, et in sub[["ts", "事件类型"]].values:
                                if et == "release":
                                    rel_ts = ts
                                elif et == "arrived" and rel_ts is not None:
                                    v = (ts - rel_ts).total_seconds()
                                    if 0 <= v < 7200:
                                        rows.append({
                                            "Station":    ws_name,
                                            "Hour":       rel_ts.strftime("%H:00"),
                                            "Release_ts": rel_ts,
                                            "Arrived_ts": ts,
                                        })
                                    rel_ts = None
                        if rows:
                            switch_sn = _sn("switch_events", p)
                            pd.DataFrame(rows).to_excel(w, sheet_name=switch_sn, index=False)
                            sws = wb[switch_sn]
                            _style_header(sws)
                            # Col E = Switch_s formula; col F = annotation
                            sws.cell(1, 5, "Switch_s")
                            sws.cell(1, 5).font = Font(bold=True, color="FFFFFF")
                            sws.cell(1, 5).fill = PatternFill("solid", fgColor=_DARK)
                            _note_cell(sws, 1, 6,
                                "Switch_s = (Arrived_ts − Release_ts) × 86400  "
                                "| Event pair: release → next arrived at same Station  "
                                "| Filter: 0 ≤ Switch_s < 7200")
                            for r in range(2, len(rows) + 2):
                                sws.cell(r, 5, f"=(D{r}-C{r})*86400")
                    except Exception:
                        pass

                # ── 6. Dwell utilisation formula chain ────────────────────
                #    One row per Station × Hour.  Every metric is a live formula.
                #    Columns:
                #      A=Station  B=Hour  C=Avg_Pick_s  D=Avg_Switch_s
                #      E=Implied_TPH_fixed(6s)  F=Implied_TPH_actual
                #      G=Actual_Completions  H=Util_fixed_%  I=Util_actual_%
                if tp_sn and pick_sn and ws_order:
                    try:
                        dwell_sn = _sn("dwell_derived", p)
                        dws      = wb.create_sheet(dwell_sn)
                        col_hdrs = [
                            "Station", "Hour",
                            "Avg_Pick_s", "Avg_Switch_s",
                            "Implied_TPH_fixed", "Implied_TPH_actual",
                            "Actual_Completions",
                            "Util_fixed_%", "Util_actual_%",
                        ]
                        for ci, h in enumerate(col_hdrs, 1):
                            dws.cell(1, ci, h)
                        _style_header(dws)

                        # Row 2 — formula legend
                        formulas_legend = [
                            "(row value)", "(row value)",
                            f"AVERAGEIFS(pick_events[Pick_s], Station, Hour)",
                            f"AVERAGEIFS(switch_events[Switch_s], Station, Hour) — '' if no switch data",
                            "3600 / (Avg_Pick_s + 6)  ← 6 s assumed fixed switch",
                            "3600 / (Avg_Pick_s + Avg_Switch_s)",
                            f"COUNTIFS(tp_events[Station], Station, tp_events[Hour], Hour)",
                            "Actual_Completions / Implied_TPH_fixed × 100",
                            "Actual_Completions / Implied_TPH_actual × 100",
                        ]
                        for ci, txt in enumerate(formulas_legend, 1):
                            _note_cell(dws, 2, ci, txt)

                        hours = [f"{h:02d}:00" for h in range(24)]
                        row   = 3
                        for st in ws_order:
                            for hr in hours:
                                dws.cell(row, 1, st)
                                dws.cell(row, 2, hr)

                                # C — avg pick time
                                dws.cell(row, 3,
                                    f"=IFERROR(AVERAGEIFS('{pick_sn}'!$F:$F,"
                                    f"'{pick_sn}'!$B:$B,$A{row},"
                                    f"'{pick_sn}'!$C:$C,$B{row}),\"\")")

                                # D — avg switch time (blank if no switch sheet)
                                if switch_sn:
                                    dws.cell(row, 4,
                                        f"=IFERROR(AVERAGEIFS('{switch_sn}'!$E:$E,"
                                        f"'{switch_sn}'!$A:$A,$A{row},"
                                        f"'{switch_sn}'!$B:$B,$B{row}),\"\")")

                                # E — implied TPH with fixed 6 s switch
                                dws.cell(row, 5, f"=IFERROR(3600/(C{row}+6),\"\")")

                                # F — implied TPH with actual switch
                                if switch_sn:
                                    dws.cell(row, 6,
                                        f"=IFERROR(3600/(C{row}+D{row}),\"\")")

                                # G — actual completions
                                dws.cell(row, 7,
                                    f"=COUNTIFS('{tp_sn}'!$C:$C,$A{row},"
                                    f"'{tp_sn}'!$B:$B,$B{row})")

                                # H — utilisation % (fixed switch)
                                dws.cell(row, 8, f"=IFERROR(G{row}/E{row}*100,\"\")")

                                # I — utilisation % (actual switch)
                                if switch_sn:
                                    dws.cell(row, 9,
                                        f"=IFERROR(G{row}/F{row}*100,\"\")")

                                row += 1
                    except Exception:
                        pass

                # ── 7. Cycle time events with formula Cycle_min column ────
                if tlc is not None:
                    try:
                        tj = tlc.dropna(subset=[TOTAL_DURATION_COL]).copy()
                        tj = tj[(tj[TOTAL_DURATION_COL] >= 0) & (tj[TOTAL_DURATION_COL] < 7200)]
                        cycle_sn = _sn("cycle_events", p)
                        tj.to_excel(w, sheet_name=cycle_sn, index=False)
                        cws = wb[cycle_sn]
                        _style_header(cws)

                        # Locate the duration column position (1-based)
                        dur_idx  = list(tj.columns).index(TOTAL_DURATION_COL) + 1
                        dur_col  = get_column_letter(dur_idx)
                        next_col = len(tj.columns) + 1

                        cws.cell(1, next_col, "Cycle_min")
                        cws.cell(1, next_col).font = Font(bold=True, color="FFFFFF")
                        cws.cell(1, next_col).fill = PatternFill("solid", fgColor=_DARK)
                        _note_cell(cws, 1, next_col + 1,
                            f"Cycle_min = {TOTAL_DURATION_COL} ÷ 60  "
                            f"| Outliers filtered: 0 ≤ duration < 7200 s")
                        for r in range(2, len(tj) + 2):
                            cws.cell(r, next_col, f"={dur_col}{r}/60")

                        # Cycle stats (computed)
                        stats = (
                            tj[TOTAL_DURATION_COL].div(60)
                            .describe(percentiles=[0.5, 0.75, 0.9, 0.95, 0.99])
                            .round(2)
                            .reset_index()
                        )
                        stats.columns = ["Metric", "Cycle_min"]
                        stats_sn = _sn("cycle_stats", p)
                        stats.to_excel(w, sheet_name=stats_sn, index=False)
                        _style_header(wb[stats_sn])
                        _method_cell(wb[stats_sn], 1, 3,
                            "Computed: percentiles of (Total_Duration_s / 60). "
                            "Rows with duration ≥ 7200 s or NaN are excluded before calculation.")
                    except Exception:
                        pass

                # ── 8. Delivery leg ───────────────────────────────────────
                if tlc is not None:
                    try:
                        dest_col = next((c for c in tlc.columns if "目标位置" in str(c)), None)
                        leg_col  = next((c for c in tlc.columns if "完成耗时" in str(c)), None)
                        if dest_col and leg_col:
                            df_leg = tlc[tlc[dest_col].astype(str).str.startswith("LABOR")].copy()
                            df_leg = df_leg.dropna(subset=[leg_col])
                            df_leg = df_leg[(df_leg[leg_col] > 0) & (df_leg[leg_col] < 3600)]
                            if not df_leg.empty:
                                df_leg["Station"]          = df_leg[dest_col].astype(str)
                                df_leg["Delivery_Leg_s"]   = df_leg[leg_col].astype(float)
                                leg_sn = _sn("delivery_leg", p)
                                df_leg[["Station", "Delivery_Leg_s"]].to_excel(
                                    w, sheet_name=leg_sn, index=False)
                                lws = wb[leg_sn]
                                _style_header(lws)
                                _note_cell(lws, 1, 3,
                                    f"Delivery_Leg_s = '{leg_col}' column from lifecycle sheet  "
                                    f"| Filter: destination starts with 'LABOR', 0 < value < 3600")

                                ws_order_l = cfg.get("ws_order", sorted(df_leg["Station"].unique()))
                                stats_leg  = (
                                    df_leg.groupby("Station")["Delivery_Leg_s"]
                                    .describe(percentiles=[0.25, 0.5, 0.75, 0.9])
                                    .round(1)
                                    .reindex(ws_order_l)
                                )
                                leg_stats_sn = _sn("leg_stats", p)
                                stats_leg.to_excel(w, sheet_name=leg_stats_sn)
                                _style_header(wb[leg_stats_sn])
                                _method_cell(wb[leg_stats_sn], 1, len(stats_leg.columns) + 2,
                                    "Computed: pandas .describe() on Delivery_Leg_s grouped by Station.")
                    except Exception:
                        pass

                # ── 9. Fleet utilisation (computed — interval logic) ───────
                if lsr is not None:
                    try:
                        lsr_u   = lsr.copy()
                        lsr_u["ts"] = pd.to_datetime(lsr_u["时间戳"], errors="coerce")
                        rb_col  = next((c for c in lsr_u.columns if "机器人编号" in str(c)), None)
                        rbt_col = next((c for c in lsr_u.columns if "机器人类型" in str(c)), None)
                        if rb_col and rbt_col:
                            lsr_u = lsr_u.dropna(subset=["ts", rb_col, rbt_col])
                            util_rows: list[dict] = []
                            for rtype, fleet_df in lsr_u.groupby(rbt_col):
                                fleet_size = fleet_df[rb_col].nunique()
                                fd = fleet_df.copy()
                                fd["hour"] = fd["ts"].dt.floor("h")
                                for hr, cnt in fd.groupby("hour")[rb_col].nunique().items():
                                    util_rows.append({
                                        "Hour":          hr.strftime("%H:00"),
                                        "Robot_Type":    rtype,
                                        "Active_Robots": int(cnt),
                                        "Fleet_Size":    fleet_size,
                                        "Util_pct":      round(cnt / fleet_size * 100, 1) if fleet_size else None,
                                    })
                            if util_rows:
                                fleet_sn = _sn("fleet_util", p)
                                fleet_df_out = pd.DataFrame(util_rows)
                                fleet_df_out.to_excel(w, sheet_name=fleet_sn, index=False)
                                fws = wb[fleet_sn]
                                _style_header(fws)
                                _method_cell(fws, 1, len(fleet_df_out.columns) + 2,
                                    "Computed: Active_Robots = unique robot IDs seen in each hour-bucket.  "
                                    "Fleet_Size = total unique robots of that type across the whole day.  "
                                    "Util_pct = Active_Robots / Fleet_Size × 100.  "
                                    "Note: real-time concurrent-delivery profile is in the HTML report "
                                    "(requires O(n²) interval overlap — not reproducible as a cell formula).")
                    except Exception:
                        pass

            # Remove the default empty sheet openpyxl creates for new workbooks
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        return out_path

    # ── done ─────────────────────────────────────────────────────────────────

    def _on_done(self, html_path: str):
        self._running   = False
        self._html_path = html_path
        self._btn_rep.configure(state="normal")
        self._btn_dir.configure(state="normal")
        self._log_write(f"  Report → {os.path.basename(html_path)}\n", "muted")
        self._set_status(f"Complete — {self._outdir}")
        # Re-evaluate run button in case new files arrived during the run
        self._maybe_enable_run()

    def _open_report(self):
        if self._html_path and os.path.isfile(self._html_path):
            os.startfile(self._html_path)

    def _open_folder(self):
        if self._outdir and os.path.isdir(self._outdir):
            os.startfile(self._outdir)

    def _open_combined(self):
        if self._combined_path and os.path.isfile(self._combined_path):
            os.startfile(self._combined_path)

    def run(self):
        self.root.mainloop()


# ══════════════════════════════════════════════════════════════════════════════
# HEADLESS CLI MODE
# ══════════════════════════════════════════════════════════════════════════════

def _run_headless(path: str):
    outdir = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                          "asrs_analysis_output")
    os.makedirs(outdir, exist_ok=True)

    # ── Validate file ────────────────────────────────────────────────────
    file_vr = validate_file_path(path)
    for w in file_vr.warnings:
        print(f"  [warn] {w}")
    if not file_vr.ok:
        for e in file_vr.errors:
            print(f"  [ERROR] {e}")
        sys.exit(1)

    print(f"Reading: {path}\nOutput : {outdir}\n")

    data = load_data(path)

    # ── Validate data quality ────────────────────────────────────────────
    data_vr = validate_data(data)
    for w in data_vr.warnings:
        print(f"  [warn] {w}")
    if not data_vr.ok:
        for e in data_vr.errors:
            print(f"  [ERROR] {e}")
        sys.exit(1)

    # ── Load and validate user config ────────────────────────────────────
    user_cfg, cfg_vr = load_user_config(path)
    for w in cfg_vr.warnings:
        print(f"  [warn] {w}")
    if not cfg_vr.ok:
        for e in cfg_vr.errors:
            print(f"  [ERROR] {e}")
        sys.exit(1)

    cfg      = build_config(data, user_cfg)
    registry: list[dict] = []

    for key, mod in _PIPELINE:
        if key in ("cycle", "retrieval") and data.get("lifecycle") is None:
            print(f"  [skip] {key} — no lifecycle sheet")
            continue
        try:
            charts = mod.run(data, cfg)
            registry.extend(charts)
            print(f"  [done] {key}  ({len(charts)} chart(s))")
        except Exception:
            print(f"  [FAIL] {key}")
            traceback.print_exc()

    day_label = os.path.splitext(os.path.basename(path))[0]
    html = generate_html_report(outdir, [{"label": day_label, "registry": registry}])
    print(f"\nComplete.\n  {outdir}\n  {html}")


# ══════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def main():
    if len(sys.argv) > 1 and os.path.isfile(sys.argv[1]):
        _run_headless(sys.argv[1])
    else:
        AnalyzerApp().run()


if __name__ == "__main__":
    main()
